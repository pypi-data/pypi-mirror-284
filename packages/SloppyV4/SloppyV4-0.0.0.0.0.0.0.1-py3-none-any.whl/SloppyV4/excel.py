#!/usr/bin/env python3
# -*- coding:utf-8 -*-
'''
# File: excel.py
# Project: SloppyV4
# Created Date: 2024-07-15, 03:28:15
# Author: Chungman Kim(h2noda@unipark.kr)
# Last Modified: Mon Jul 15 2024
# Modified By: Chungman Kim
# Copyright (c) 2024 Unipark
# HISTORY:
# Date      	By	Comments
# ----------	---	----------------------------------------------------------
'''

import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font, numbers
from openpyxl import Workbook


class Xlsx:
    """ Class : Xlsx
    """
    def __init__(self, filename, sheetname="sheet", idx=0, mode="n"):
        """Mode = n : New Excel File
                  r : Read Excel File

        Args:
            filename (String): Excel File Name
            sheetname (String): Excel Sheet Name
            mode (str, optional): Excel file Mode, Defaults to "n".
        """
        if (mode == "n"):
            self.filename = filename
            self.sheetname = sheetname
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = sheetname
        elif (mode == "r"):
            self.filename = filename
            self.wsname = sheetname
            self.wb = openpyxl.load_workbook(filename)
            ws_name = self.wb.sheetnames()
            if sheetname in ws_name:
                self.ws = self.wb[ws_name[0]]
            elif sheetname == "":
                self.ws = self.wb[ws_name[idx]]
            else:
                self.ws = self.wb[sheetname]
