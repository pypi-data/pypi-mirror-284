import matplotlib.transforms as mtransforms
import numpy as np
import scipy.linalg
from .consts import *
import matplotlib.pyplot as plt
from matplotlib.axes import Axes
from matplotlib.figure import Figure
from matplotlib.lines import Line2D
from ..filesop.filesave import FilesSave
import openpyxl
from scipy.optimize import curve_fit

plt.rc("font", family="Times New Roman")  # change the font of plot
plt.rcParams["mathtext.fontset"] = "stix"


class PubMethod:
    def __init__(self) -> None:
        pass

    @staticmethod
    def add_right_cax(ax: Axes, pad, width) -> Axes:  # ax is the axe of figure
        axpos = ax.get_position()
        caxpos = mtransforms.Bbox.from_extents(
            axpos.x1 + pad, axpos.y0, axpos.x1 + pad + width, axpos.y1
        )
        cax = ax.figure.add_axes(caxpos)

        return cax

    @staticmethod
    def sym_limit(mat_in: np.ndarray, split_realimag=False):
        if split_realimag:
            real_vmin = np.min(np.real(mat_in))
            real_vmax = np.max(np.real(mat_in))
            imag_vmin = np.min(np.imag(mat_in))
            imag_vmax = np.max(np.imag(mat_in))
            real_lim = (
                abs(real_vmin) if abs(real_vmin) > abs(real_vmax) else abs(real_vmax)
            )
            imag_lim = (
                abs(imag_vmin) if abs(imag_vmin) > abs(imag_vmax) else abs(imag_vmax)
            )
            return real_lim, imag_lim
        else:
            vmin = np.min(mat_in)
            vmax = np.max(mat_in)
            limit = abs(vmin) if abs(vmin) > abs(vmax) else abs(vmax)
            return limit

    @staticmethod
    def cmap_scatters(
        x_list,
        y_list,
        c_arr,
        bound_vecs,
        saveInst: FilesSave,
        fname,
        marker_size=5.5,
        marker_type="h",
        vmin=None,
        vmax=None,
        title="",
        cmap="bwr",
        colorbar_pad=0.01,
        colorbar_width=0.01,
        clabel_name="",
        titlesize=14,
        fontsize=12,
        figax_in=None,
    ):
        if figax_in is None:
            fig, ax_2d = plt.subplots(figsize=(7, 7))
        else:
            fig, ax_2d = figax_in
        im_var = ax_2d.scatter(
            x_list,
            y_list,
            c=np.real(c_arr),
            s=marker_size,
            vmin=vmin,
            vmax=vmax,
            marker=marker_type,
            cmap=cmap,
        )
        ax_2d.plot(bound_vecs[:, 0], bound_vecs[:, 1], "k-")
        ax_2d.set_aspect("equal")
        ax_2d.axis("off")
        ax_2d.set_title(title, fontsize=titlesize)
        c_ax = PubMethod.add_right_cax(ax_2d, colorbar_pad, colorbar_width)
        cbar = fig.colorbar(im_var, cax=c_ax)
        cbar.set_label(
            label=clabel_name,
            fontsize=fontsize,
            labelpad=13,
        )
        saveInst.save_fig(fig, fname)
        if figax_in is None:
            plt.close(fig)
        else:
            fig.delaxes(c_ax)
            ax_2d.cla()

    @staticmethod
    def r_mat(angle):
        theta = angle / 180 * pi
        r_mat = np.array(
            [[np.cos(theta), -np.sin(theta)], [np.sin(theta), np.cos(theta)]]
        )

        return r_mat


class UnitsConversion:
    def __init__(self) -> None:
        pass

    @staticmethod
    def energy2omega(energy_in_meV):
        omega_arr = np.array(energy_in_meV) / (h_bar_eV * eV2meV)

        return omega_arr

    @staticmethod
    def wls2omega(wls_in_nm):
        energy_in_meV = 1240 / wls_in_nm * 1000
        omega_arr = np.array(energy_in_meV) / (h_bar_eV * eV2meV)

        return omega_arr


class HaCo:
    def __init__(self) -> None:
        pass

    @staticmethod
    def DiracH(p_arr: np.ndarray, angle_in):
        theta_in = angle_in / 180 * pi

        def ele_h(p_x, p_y):
            return np.array(
                [
                    [0, np.exp(-1j * theta_in) * (p_x - 1j * p_y)],
                    [np.exp(1j * theta_in) * (p_x + 1j * p_y), 0],
                ]
            )

        if len(p_arr.shape) > 1:
            "Multiple k array to form block diagonal"

            out_list = [
                ele_h(p_arr[:, 0][i], p_arr[:, 1][i]) for i in range(len(p_arr))
            ]

            return scipy.linalg.block_diag(*out_list)
        return ele_h(p_arr[0], p_arr[1])

    @staticmethod
    def ParabH(p_arr: np.ndarray, angle_in, epsilon, tperp=357):
        theta_in = angle_in / 180 * pi

        def ele_h(p_x, p_y):
            return np.array(
                [
                    [0, np.exp(-1j * theta_in) * (p_x - 1j * p_y), 0, tperp / epsilon],
                    [np.exp(1j * theta_in) * (p_x + 1j * p_y), 0, 0, 0],
                    [0, 0, 0, np.exp(-1j * theta_in) * (p_x - 1j * p_y)],
                    [tperp / epsilon, 0, np.exp(1j * theta_in) * (p_x + 1j * p_y), 0],
                ]
            )

        if len(p_arr.shape) > 1:
            "Multiple k array to form block diagonal"

            out_list = [
                ele_h(p_arr[:, 0][i], p_arr[:, 1][i]) for i in range(len(p_arr))
            ]

            return scipy.linalg.block_diag(*out_list)

        return ele_h(p_arr[0], p_arr[1])

    @staticmethod
    def AAstackedH(p_arr: np.ndarray, angle_in, epsilon, tperp_AA=93, tperp_BB=105):
        theta_in = angle_in / 180 * pi

        def ele_h(p_x, p_y):
            return np.array(
                [
                    [
                        0,
                        np.exp(-1j * theta_in) * (p_x - 1j * p_y),
                        tperp_AA / epsilon,
                        0,
                    ],
                    [
                        np.exp(1j * theta_in) * (p_x + 1j * p_y),
                        0,
                        0,
                        tperp_BB / epsilon,
                    ],
                    [
                        tperp_AA / epsilon,
                        0,
                        0,
                        np.exp(-1j * theta_in) * (p_x - 1j * p_y),
                    ],
                    [
                        0,
                        tperp_BB / epsilon,
                        np.exp(1j * theta_in) * (p_x + 1j * p_y),
                        0,
                    ],
                ]
            )

        if len(p_arr.shape) > 1:
            "Multiple k array to form block diagonal"

            out_list = [
                ele_h(p_arr[:, 0][i], p_arr[:, 1][i]) for i in range(len(p_arr))
            ]

            return scipy.linalg.block_diag(*out_list)

        return ele_h(p_arr[0], p_arr[1])


class Line:
    def __init__(self, x, y, fInst: FilesSave = FilesSave("Plots/Lines")) -> None:
        self.x, self.y = x, y
        self.fInst = fInst
        pass

    def ax_return(self, ax_return: bool, ax_fig: list[Axes, Figure]):
        if ax_return:
            return ax_fig
        else:
            plt.close(ax_fig[1])

    def plot(self, fname, xlabel="", ylabel="", title="", xlim=None):
        fig, ax = plt.subplots()
        ax.plot(self.x, self.y)
        ax.set_aspect("auto")
        ax.set_xlabel(xlabel, fontsize=12)
        ax.set_ylabel(ylabel, fontsize=12)
        ax.set_title(title, fontsize=14)
        if xlim is None:
            ax.set_xlim(ax.get_xlim())
        else:
            ax.set_xlim(xlim)
        ax.set_ylim(ax.get_ylim())
        self.fInst.save_fig(fig, fname=fname)
        plt.close()

    def multiplot(
        self,
        fname,
        legends,
        xlabel="",
        ylabel="",
        title="",
        xlim=None,
        linestyles=None,
        ax_return=False,
    ):
        fig, ax = plt.subplots()
        if linestyles is None:
            linestyles = ["-"] * len(self.x)
        else:
            pass
        for i in range(len(self.x)):
            ax.plot(self.x[i], self.y[i], linestyles[i], label=legends[i])

        ax.set_aspect("auto")
        ax.set_xlabel(xlabel, fontsize=12)
        ax.set_ylabel(ylabel, fontsize=12)
        ax.set_title(title, fontsize=14)
        if xlim is None:
            ax.set_xlim(ax.get_xlim())
        else:
            ax.set_xlim(xlim)
        ax.legend()
        ax.set_ylim(ax.get_ylim())
        self.fInst.save_fig(fig, fname=fname)
        return self.ax_return(ax_return, [ax, fig])

    def semilogy(self, fname, legends, xlabel="", ylabel="", title="", xlim=None):
        fig, ax = plt.subplots()
        for i in range(len(self.x)):
            ax.semilogy(self.x[i], self.y[i], label=legends[i])
        ax.set_aspect("auto")
        ax.set_xlabel(xlabel, fontsize=12)
        ax.set_ylabel(ylabel, fontsize=12)
        ax.set_title(title, fontsize=14)
        if xlim is None:
            ax.set_xlim(ax.get_xlim())
        else:
            ax.set_xlim(xlim)
        ax.legend()
        ax.set_ylim(ax.get_ylim())
        self.fInst.save_fig(fig, fname=fname)
        plt.close()
        return

    def center_of_curve(self):
        """
        Get the center of the 2D-curve numerically.

        #   Return
        xbar, ybar
        """
        diff_y = np.diff(self.y)
        diff_x = np.diff(self.x)

        der_y = diff_y / diff_x

        ds_factor = np.sqrt(1 + der_y**2)

        x_center_kernel = self.x[:-1] * ds_factor
        y_center_kernel = self.y[:-1] * ds_factor

        integral_s = np.trapz(ds_factor, self.x[:-1])

        integral_x = np.trapz(x_center_kernel, self.x[:-1])
        integral_y = np.trapz(y_center_kernel, self.x[:-1])

        x_bar = integral_x / integral_s
        y_bar = integral_y / integral_s

        return x_bar, y_bar

    def kb_of_curve(self) -> np.ndarray:
        xbar, ybar = self.center_of_curve()
        k, b = np.polyfit(self.x, self.y, 1)

        def func_fixp(x, k):

            return k * (x - xbar) + ybar

        opt_k = curve_fit(func_fixp, self.x, self.y, k)[0]

        return opt_k, b


class DefinedFuncs:
    def __init__(self) -> None:
        pass

    @staticmethod
    def deltaF_arct(x: np.ndarray, center=0, a=1):
        if isinstance(center, np.ndarray):
            dims = [1] * len(x.shape)
            xe = np.kron(np.ones((len(center), *dims)), x)
            center_e: np.ndarray = np.kron(
                center.reshape((len(center), *dims)), np.ones(x.shape)
            )
            return a / (np.pi * ((xe - center_e) ** 2 + a**2))
        return a / (np.pi * ((x - center) ** 2 + a**2))


class ExcelMethod:
    def __init__(self, file_path) -> None:
        self.file_path = file_path
        pass

    @staticmethod
    def excel_col_indices(len_of_column):
        """
        Create list of column indices.
        """
        basic_list = [chr(i).capitalize() for i in range(97, 123)]
        basic_len = len(basic_list)
        if len_of_column <= basic_len:
            return basic_list[:len_of_column]
        else:
            first_letter_i = len_of_column // basic_len
            secon_letter_i = len_of_column % basic_len

            extend_list = basic_list[:]
            begin_first = 0
            while first_letter_i > 0:
                first_letter_i = first_letter_i - 1
                ele_first_letter = basic_list[begin_first]
                begin_first = begin_first + 1
                if first_letter_i == 0:
                    extend_list = extend_list + [
                        ele_first_letter + ele_secon
                        for ele_secon in basic_list[:secon_letter_i]
                    ]
                else:
                    extend_list = extend_list + [
                        ele_first_letter + ele_secon for ele_secon in basic_list
                    ]
            return extend_list

    def read_xlsx_data(self, exclude_rows_num=0):
        data = openpyxl.load_workbook(self.file_path)
        outsheets = []
        for elesheet in data.sheetnames:
            sheet = data[elesheet]
            column_to_read = sheet.max_column
            max_row_num = sheet.max_row
            ele_is_column_list = []
            for column_i in range(column_to_read):
                column_tag = self.excel_col_indices(column_to_read)
                ele_column_cell = sheet[
                    "{0}{2}:{0}{1}".format(
                        column_tag[column_i], max_row_num, 1 + exclude_rows_num
                    )
                ]
                ele_cell_dat = []
                for ele_i in range(max_row_num - exclude_rows_num):
                    value = ele_column_cell[ele_i][0].value
                    if isinstance(value, str) and value[-1] == "j":
                        ele_cell_dat.append(complex(value))
                    elif isinstance(value, (float, int)):
                        ele_cell_dat.append(value)
                # ele_cell_dat = [
                #     ele_column_cell[ele_i][0].value
                #     for ele_i in range(max_row_num - exclude_rows_num)
                # ]
                ele_is_column_list.append(ele_cell_dat)
            outsheets.append(ele_is_column_list)
        return outsheets


def main():
    # x1 = np.linspace(0, 3, 100)
    # y1 = np.sin(x1)
    # y2 = np.cos(x1)
    # Line([x1, x1], [y1, y2]).multiplot(
    #     "test_linestyle", legends=["", ""], linestyles=["--", "-"]
    # )

    pass


if __name__ == "__main__":
    main()
