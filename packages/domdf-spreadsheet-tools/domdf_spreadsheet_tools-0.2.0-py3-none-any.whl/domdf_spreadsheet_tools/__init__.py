#!/usr/bin/env python
#
#  domdf_spreadsheet_tools.py
"""
Tools for creating and formatting spreadsheets with Python and OpenPyXL.
"""
#
#  Copyright 2018-2020 Dominic Davis-Foster <dominic@davis-foster.co.uk>
#
#  This program is free software; you can redistribute it and/or modify
#  it under the terms of the GNU Lesser General Public License as published by
#  the Free Software Foundation; either version 3 of the License, or
#  (at your option) any later version.
#
#  This program is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
#  GNU Lesser General Public License for more details.
#
#  You should have received a copy of the GNU Lesser General Public License
#  along with this program; if not, write to the Free Software
#  Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston,
#  MA 02110-1301, USA.
#

# stdlib
import csv
import locale
import os
import traceback
from typing import Any, Dict, List, Mapping, Optional, Union

# 3rd party
from domdf_python_tools.typing import PathLike
from openpyxl import Workbook, load_workbook  # type: ignore[import]
from openpyxl.styles import Alignment  # type: ignore[import]
from openpyxl.utils import get_column_letter  # type: ignore[import]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import]

__author__ = "Dominic Davis-Foster"
__copyright__ = "Copyright 2018-2020 Dominic Davis-Foster"
__license__ = "LGPLv3+"
__version__ = "0.1.6"
__email__ = "dominic@davis-foster.co.uk"

__all__ = ["append_to_xlsx", "format_header", "format_sheet", "make_column_property_dict", "col_width_from_cm"]


def append_to_xlsx(
		csv_input_file: PathLike,
		xlsx_output_file: PathLike,
		sheet_title: Optional[str] = None,
		separator: str = ',',
		overwrite: bool = False,
		to_floats: bool = False,
		encoding: str = "UTF-8"
		) -> None:
	"""
	Add CSV file to xlsx file as a new worksheet.

	:param csv_input_file: filepath of CSV file.
	:param xlsx_output_file: filepath of xlsx file.
	:param sheet_title: Title of sheet to append. Default is the value of ``csv_input_file``.
	:no-default sheet_title:
	:param separator: Separator for reading CSV file.
	:param overwrite: Whether to overwrite the xlsx output file
		(i.e. create a new file containing just the new sheet).
	:param to_floats: Whether to read strings with thousand separators as floats.
	:param encoding: The encoding to read the file as.
	"""

	# Setup for reading strings with thousand separators as floats
	# From https://stackoverflow.com/a/31074271
	locale.setlocale(locale.LC_ALL, '')

	if sheet_title is None:
		sheet_title = os.path.splitext(os.path.basename(csv_input_file))[0]

	if overwrite:
		wb = Workbook()
		ws = wb.active
		wb.remove_sheet(ws)
	else:
		wb = load_workbook(xlsx_output_file)

	wb.create_sheet(sheet_title)
	ws = wb[sheet_title]

	with open(csv_input_file, encoding=encoding) as f:
		reader = csv.reader(f, delimiter=separator)

	for row in reader:
		try:
			if to_floats:
				row_buffer: List[Union[str, float]] = []
				for cell in row:
					try:
						row_buffer.append(locale.atof(cell))
					except BaseException:
						row_buffer.append(cell)
				ws.append(row_buffer)
			else:
				ws.append(row)
		except BaseException:
			traceback.print_exc()  # print the error
			print(row)

	wb.save(xlsx_output_file)


def format_sheet(
		ws: Worksheet,
		number_format_list: Optional[Dict[str, str]] = None,
		width_list: Optional[Dict[str, float]] = None,
		alignment_list: Optional[Dict[str, str]] = None,
		) -> None:
	"""
	Format columns of an xlsx worksheet.

	:param ws: The worksheet to format.
	:type ws: :class:`openpyxl.worksheet.worksheet.Worksheet`.
	:param number_format_list: dictionary of number format strings for each column letter.
	:param width_list: dictionary of widths for each column letter.
	:param alignment_list: dictionary of alignments (``left``, ``right``, or ``center``) for each column letter.
	"""

	# for row in ws.iter_rows("A1:{}{}".format(get_column_letter(ws.max_column), ws.max_row)):
	for row in ws[f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"]:
		for cell in row:
			cell.alignment = Alignment(vertical="center", wrap_text=False)

	if number_format_list:
		for column in number_format_list:
			# for row in ws.iter_rows('{0}{1}:{0}{2}'.format(column, 3, ws.max_row)):
			for row in ws[f"{column}{3}:{column}{ws.max_row}"]:
				for cell in row:
					cell.number_format = number_format_list[column]

	for column_cells in ws.columns:
		length = max(len(str(cell.value)) for cell in column_cells)
		if length < 1:
			length = 1
		ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length
	# ws.column_dimensions[column_cells[0].column].bestFit = True

	if width_list:
		for column in width_list:
			if width_list[column] == 0:
				ws.column_dimensions[column].hidden = True
			else:
				ws.column_dimensions[column].width = width_list[column]

	if alignment_list:
		for column in alignment_list:
			# for row in ws.iter_rows("{0}{1}:{0}{2}".format(column, ws.min_row, ws.max_row)):
			for row in ws[f"{column}{ws.min_row}:{column}{ws.max_row}"]:
				for cell in row:
					cell.alignment = Alignment(
							horizontal=alignment_list[column],
							vertical="center",
							wrap_text=False,
							)


def format_header(
		ws: Worksheet,
		alignment_list: Dict[str, str],
		start_row: int = 1,
		end_row: int = 1,
		) -> None:
	"""
	Format the alignment of the header rows of a worksheet.

	:param ws: The worksheet to format.
	:type ws: :class:`openpyxl.worksheet.worksheet.Worksheet`.
	:param alignment_list: dictionary of alignments (left, right, center) for each column letter.
	:param start_row: The row to start formatting on.
	:param end_row: The row to end formatting on.
	"""

	for column in alignment_list:
		# for row in ws.iter_rows("{0}{1}:{0}{2}".format(column, start_row, end_row)):
		for row in ws[f"{column}{start_row}:{column}{end_row}"]:
			for cell in row:
				cell.alignment = Alignment(horizontal=alignment_list[column], vertical="center", wrap_text=False)


def make_column_property_dict(
		indict: Mapping,
		outdict: Optional[Dict[str, Any]] = None,
		offset_dict: Optional[Mapping] = None,
		repeat: int = 1,
		length: int = 1,
		) -> Dict[str, Any]:
	"""
	Generate property lists from integer values.

	:param indict: Property values to add to the property dict.
	:param outdict: Dictionary of properties for each column letter.
	:param offset_dict:
	:param repeat:
	:param length:

	.. TODO:: Finish this docstring; check usage in GunShotMatch
	"""

	if outdict is None:
		outdict = {}

	for index in indict:
		for offset in range(repeat):
			outdict[get_column_letter(int(index) + (length * offset))] = indict[index]

	if offset_dict:
		offset = repeat * length
		for index in offset_dict:
			outdict[get_column_letter(int(index) + offset)] = offset_dict[index]

	return outdict


def col_width_from_cm(width: float) -> float:
	"""
	Returns an estimate of the width of the column in pixels, given the size in centimeters.

	:param width: The width in centimeters.

	.. versionadded:: 0.2.0
	"""

	# https://docs.microsoft.com/en-us/office/trouleshoot/excel/determine-column-widths

	# width in px / 38 = width in cm
	# width in cm * 38 = width in px
	# 10 = 75px ~= 1.97cm

	return width / (1.97 / 10)
